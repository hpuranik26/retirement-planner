"""
Retirement Withdrawal Planner
Usage:
  python3 build_retirement_planner.py [scenario.xlsx]

  First run on a new file → creates input template with defaults, exits.
  Edit amber cells in the Inputs tab, then run again to produce results.
  Subsequent runs → reads Inputs + Tax Law year, writes all output tabs.
  Tax Tables tab shows the exact law values used in this run.
"""

import sys, os, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# TAX LAW DATABASE  — add a new year entry each time IRS/CMS publishes updates
# Rates (10/12/22% etc.) are statutory and don't change between entries.
# Thresholds and amounts are what change year to year.
# ═══════════════════════════════════════════════════════════════════════════════
TAX_LAW_DB = {
    # ── 2025 ─────────────────────────────────────────────────────────────────
    # Sources: IRS Rev. Proc. 2024-40 (federal/LTCG), CMS 2025 announcement
    #          (IRMAA/Medicare), Oregon DOR 2025 withholding tables
    2025: {
        # Federal MFJ brackets (IRS Rev. Proc. 2024-40)
        'fed_brackets': [
            (0,        23_200,  0.10),
            (23_200,   94_300,  0.12),
            (94_300,  201_050,  0.22),
            (201_050, 383_900,  0.24),
            (383_900, 487_450,  0.32),
            (487_450, 731_200,  0.35),
            (731_200,  1e12,    0.37),
        ],
        'std_ded_mfj':   30_000,   # MFJ standard deduction (Rev. Proc. 2024-40)
        'senior_bonus':   1_600,   # per spouse 65+ (pre-OBBBA additional std ded)
        # Oregon MFJ brackets (OR DOR 2025)
        'or_brackets': [
            (0,        18_400,  0.0475),
            (18_400,   46_200,  0.0675),
            (46_200,  250_000,  0.0875),
            (250_000,  1e12,    0.099),
        ],
        'or_std_ded':    4_840,    # Oregon MFJ standard deduction 2025
        # LTCG / Qualified dividends MFJ (Rev. Proc. 2024-40)
        'ltcg_brackets': [
            (0,        96_700,  0.00),
            (96_700,  600_050,  0.15),
            (600_050,  1e12,    0.20),
        ],
        'niit_threshold': 250_000,  # IRC §1411 — not inflation-adjusted per statute
        # IRMAA MFJ 2025 (CMS; based on 2023 MAGI)
        # Surcharge = Part B IRMAA + Part D IRMAA per person per year (above standard)
        # Part B standard premium 2025: $185.00/mo
        'irmaa': [
            (212_000,      0.0),   # ≤ $212K: standard, no surcharge
            (266_000,   1_052.40), # $212K–$266K: ($74.00 + $13.70) × 12
            (334_000,   2_694.00), # $266K–$334K: ($185.00 + $35.40) × 12
            (402_000,   4_236.00), # $334K–$402K: ($295.90 + $57.00) × 12
            (750_000,   5_832.00), # $402K–$750K: ($406.90 + $78.60) × 12
            (1e12,      6_309.60), # ≥ $750K:     ($443.90 + $85.80) × 12
        ],
        'qcd_limit':    105_000,   # per person per year (indexed to inflation)
        'rmd_start_age':     75,   # SECURE 2.0 — born 1960+
    },

    # ── 2026 ─────────────────────────────────────────────────────────────────
    # Sources: IRS Rev. Proc. 2025-32 (federal/LTCG, taxfoundation.org confirmed),
    #          One Big Beautiful Bill Act / OBBBA (std ded, senior bonus),
    #          CMS 2026 Medicare announcement (IRMAA/Part B/D),
    #          Oregon HB 3753 + OR DOR 2026 (OR brackets & standard deduction)
    2026: {
        # Federal MFJ brackets (IRS Rev. Proc. 2025-32)
        'fed_brackets': [
            (0,        24_800,  0.10),
            (24_800,  100_800,  0.12),
            (100_800, 211_400,  0.22),
            (211_400, 403_550,  0.24),
            (403_550, 512_450,  0.32),
            (512_450, 768_700,  0.35),
            (768_700,  1e12,    0.37),
        ],
        'std_ded_mfj':   32_200,   # MFJ (OBBBA + IRS inflation adj., Rev. Proc. 2025-32)
        'senior_bonus':   6_000,   # per spouse 65+ (One Big Beautiful Bill Act)
        # Oregon MFJ brackets (OR HB 3753 + OR DOR 2026)
        'or_brackets': [
            (0,        20_400,  0.0475),
            (20_400,   51_000,  0.0675),
            (51_000,  250_000,  0.0875),
            (250_000,  1e12,    0.099),
        ],
        'or_std_ded':    9_680,    # Oregon MFJ std ded 2026 (OR HB 3753)
        # LTCG / Qualified dividends MFJ (IRS Rev. Proc. 2025-32)
        'ltcg_brackets': [
            (0,        98_900,  0.00),
            (98_900,  613_700,  0.15),
            (613_700,  1e12,    0.20),
        ],
        'niit_threshold': 250_000,  # IRC §1411 — not inflation-adjusted per statute
        # IRMAA MFJ 2026 (CMS; based on 2024 MAGI)
        # Surcharge = Part B IRMAA + Part D IRMAA per person per year (above standard)
        # Part B standard premium 2026: $202.90/mo
        'irmaa': [
            (218_000,      0.0),   # ≤ $218K: standard, no surcharge
            (274_000,   1_147.92), # $218K–$274K: ($81.16 + $14.50) × 12
            (342_000,   2_884.80), # $274K–$342K: ($202.90 + $37.50) × 12
            (410_000,   4_620.48), # $342K–$410K: ($324.64 + $60.40) × 12
            (750_000,   6_356.16), # $410K–$750K: ($446.38 + $83.30) × 12
            (1e12,      6_935.52), # ≥ $750K:     ($486.96 + $91.00) × 12
        ],
        'qcd_limit':    108_000,   # per person per year (2026 inflation-adjusted)
        'rmd_start_age':     75,   # SECURE 2.0 — born 1960+
    },
    # ── Add future years here as IRS/CMS publishes them ──────────────────────
    # 2027: { 'fed_brackets': [...], 'std_ded_mfj': ..., 'senior_bonus': ..., ... }
}

def get_tax_law(requested_year):
    """
    Return (tax_law_dict, actual_year_used, fallback_used).
    If requested year not in DB, falls back to latest available year.
    """
    if requested_year in TAX_LAW_DB:
        return TAX_LAW_DB[requested_year], requested_year, False
    latest = max(TAX_LAW_DB.keys())
    return TAX_LAW_DB[latest], latest, True

# ═══════════════════════════════════════════════════════════════════════════════
# DEFAULT INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULTS = {
    'spouse1_age':      58,
    'spouse2_age':      53,
    'spouse1_birth':    1968,
    'spouse2_birth':    1973,
    'child1_birth':     2005,   # enter 0 if no child
    'child2_birth':     2007,   # enter 0 if no child
    'child3_birth':     0,      # enter 0 if no child
    'trad_bal':         4_000_000,
    'taxable_bal':      2_800_000,
    'roth_bal':           500_000,
    'withdraw_rate':    0.06,   # Fallback if no 3-phase rates set
    'after_tax_rate':   0.0,    # If > 0: each phase re-anchors after-tax cash = portfolio × phase_rate
    'gogo_rate':        0.05,   # Go-Go phase gross withdrawal rate
    'slowgo_rate':      0.04,   # Slow-Go phase gross withdrawal rate
    'nogo_rate':        0.03,   # No-Go phase gross withdrawal rate
    'gogo_years':       0,      # Years in Go-Go phase  (0 = retire_years ÷ 3)
    'slowgo_years':     0,      # Years in Slow-Go phase (0 = retire_years ÷ 3)
    'nogo_years':       0,      # Years in No-Go phase  (0 = remainder)
    'transition_years': 3,      # Years to taper spending between phases (0 = instant step)
    'conv_bracket':     0.24,
    'retire_year':      2027,
    'retire_years':     35,
    'tax_law_year':     2026,
    'ss1_fra_mo':       4_200,
    'ss2_fra_mo':       2_100,
    'ss1_claim_age':    70,
    'ss2_claim_age':    67,
    'growth_rate':      0.06,
    'price_return':     0.04,
    'div_yield':        0.02,
    'inflation':        0.025,
    'ss_cola':          0.025,
    'aca_base_21':      528.0,
    'aca_inf':          0.05,
    'medicare_base_mo': 450.0,
    'med_inf':          0.05,
}

# ═══════════════════════════════════════════════════════════════════════════════
# INPUT SHEET DEFINITION
# ═══════════════════════════════════════════════════════════════════════════════
INPUT_ROWS = [
    ("Spouse 1 Current Age",         'spouse1_age',      'int',   "As of today",                                  "FAMILY"),
    ("Spouse 2 Current Age",         'spouse2_age',      'int',   "As of today",                                  "FAMILY"),
    ("Spouse 1 Birth Year",          'spouse1_birth',    'int',   "Used for Medicare, RMD, SS timing",            "FAMILY"),
    ("Spouse 2 Birth Year",          'spouse2_birth',    'int',   "Used for Medicare, RMD, SS timing",            "FAMILY"),
    ("Child 1 Birth Year",           'child1_birth',     'int',   "Enter 0 if no child. On parents' ACA until age 26.", "FAMILY"),
    ("Child 2 Birth Year",           'child2_birth',     'int',   "Enter 0 if no child. On parents' ACA until age 26.", "FAMILY"),
    ("Child 3 Birth Year",           'child3_birth',     'int',   "Enter 0 if no child. On parents' ACA until age 26.", "FAMILY"),
    ("Tax-Deferred Balance",         'trad_bal',         'money', "Traditional IRA / 401(k) — pre-tax",          "PORTFOLIO"),
    ("Taxable Brokerage Balance",    'taxable_bal',      'money', "Brokerage account — LTCG on gains",           "PORTFOLIO"),
    ("Tax-Free (Roth) Balance",      'roth_bal',         'money', "Roth IRA / Roth 401(k)",                      "PORTFOLIO"),
    ("Retirement Start Year",        'retire_year',      'int',   "First year of withdrawals",                   "WITHDRAWAL STRATEGY"),
    ("Retirement Length (years)",    'retire_years',     'int',   "Number of years to project",                  "WITHDRAWAL STRATEGY"),
    ("Go-Go Withdrawal Rate",        'gogo_rate',        'pct',   "Phase 1 (active early retirement) — spending re-anchors at each phase start", "WITHDRAWAL STRATEGY"),
    ("Go-Go Years",                  'gogo_years',       'int',   "Years in Go-Go phase (0 = retire_years ÷ 3)", "WITHDRAWAL STRATEGY"),
    ("Slow-Go Withdrawal Rate",      'slowgo_rate',      'pct',   "Phase 2 (mid retirement, slower pace)",       "WITHDRAWAL STRATEGY"),
    ("Slow-Go Years",                'slowgo_years',     'int',   "Years in Slow-Go phase (0 = retire_years ÷ 3)", "WITHDRAWAL STRATEGY"),
    ("No-Go Withdrawal Rate",        'nogo_rate',        'pct',   "Phase 3 (later years, mostly home/care)",     "WITHDRAWAL STRATEGY"),
    ("No-Go Years",                  'nogo_years',       'int',   "Years in No-Go phase (0 = remainder of retire_years)", "WITHDRAWAL STRATEGY"),
    ("Phase Transition Years",       'transition_years', 'int',   "Years to taper spending at each phase change (0 = instant step-down)", "WITHDRAWAL STRATEGY"),
    ("Roth Conversion Bracket",      'conv_bracket',     'pct',   "Fill Trad→Roth up to this federal bracket",   "WITHDRAWAL STRATEGY"),
    ("Tax Law Year",                 'tax_law_year',     'int',   f"Base tax law year. Available: {sorted(TAX_LAW_DB.keys())}. Future years fall back to latest.", "WITHDRAWAL STRATEGY"),
    ("Spouse 1 SS at FRA ($/mo)",    'ss1_fra_mo',       'money', "Monthly benefit at Full Retirement Age (67)", "SOCIAL SECURITY"),
    ("Spouse 2 SS at FRA ($/mo)",    'ss2_fra_mo',       'money', "Monthly benefit at Full Retirement Age (67)", "SOCIAL SECURITY"),
    ("Spouse 1 SS Claiming Age",     'ss1_claim_age',    'int',   "62–70. Delay to 70 = 124% of PIA",           "SOCIAL SECURITY"),
    ("Spouse 2 SS Claiming Age",     'ss2_claim_age',    'int',   "62–70. No spousal gain past FRA (67)",        "SOCIAL SECURITY"),
    ("Portfolio Growth Rate",        'growth_rate',      'pct',   "Annual return — all accounts",                "RETURNS & INFLATION"),
    ("Taxable Price Return",         'price_return',     'pct',   "Taxable account price appreciation only",     "RETURNS & INFLATION"),
    ("Taxable Dividend Yield",       'div_yield',        'pct',   "Qualified dividends — withdrawn each year",   "RETURNS & INFLATION"),
    ("Inflation Rate",               'inflation',        'pct',   "Applied to brackets, deductions, SS COLA",    "RETURNS & INFLATION"),
    ("Social Security COLA",         'ss_cola',          'pct',   "Annual SS cost-of-living adjustment",         "RETURNS & INFLATION"),
    ("ACA Base Rate (age 21, $/mo)", 'aca_base_21',      'money', "Gold plan, Portland OR — adjust for area",    "HEALTH INSURANCE"),
    ("ACA Annual Inflation",         'aca_inf',          'pct',   "Health insurance premium inflation",          "HEALTH INSURANCE"),
    ("Medicare Base ($/person/mo)",  'medicare_base_mo', 'money', "Part B + Part D + Medigap G per person",      "HEALTH INSURANCE"),
    ("Medicare Annual Inflation",    'med_inf',          'pct',   "Medicare premium inflation",                  "HEALTH INSURANCE"),
]

# ═══════════════════════════════════════════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════════════════════════════════════════
AMBER    = "FFB300";  BLUE_H  = "1565C0";  TEAL_H  = "00695C"
GREEN_H  = "2E7D32";  GREY_H  = "546E7A";  RED_H   = "B71C1C"
PURPLE_H = "6A1B9A";  BROWN_H = "5D4037";  WHITE   = "FFFFFF"
ORANGE_H = "E65100"

SECTION_COLORS = {
    "FAMILY":              BLUE_H,  "PORTFOLIO":          TEAL_H,
    "WITHDRAWAL STRATEGY": GREEN_H, "SOCIAL SECURITY":    RED_H,
    "RETURNS & INFLATION": GREY_H,  "HEALTH INSURANCE":   "00838F",
}

def hfill(c):  return PatternFill("solid", fgColor=c)
def cfill(c):  return PatternFill("solid", fgColor=c)
def tborder():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ═══════════════════════════════════════════════════════════════════════════════
# INPUTS SHEET — create template
# ═══════════════════════════════════════════════════════════════════════════════
def create_inputs_sheet(wb, values=None):
    if "Inputs" in wb.sheetnames: del wb["Inputs"]
    ws = wb.create_sheet("Inputs", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE_H

    t = ws.cell(1, 1, "RETIREMENT WITHDRAWAL PLANNER — INPUT PARAMETERS")
    t.font = Font(bold=True, size=14, color=BLUE_H)
    ws.merge_cells("A1:D1")

    ws.cell(2, 1, "Edit the AMBER cells, save the file, then run:  python3 build_retirement_planner.py  <this_file.xlsx>")
    ws.cell(2, 1).font = Font(italic=True, size=10, color="555555")
    ws.merge_cells("A2:D2")

    for ci, (hdr, w) in enumerate([("Parameter", 38), ("Value", 18), ("Unit / Notes", 56)], 1):
        c = ws.cell(3, ci, hdr)
        c.fill = hfill(BLUE_H); c.font = Font(bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    row = 4; current_section = None
    for label, key, vtype, notes, section in INPUT_ROWS:
        if section != current_section:
            current_section = section
            color = SECTION_COLORS.get(section, GREY_H)
            c = ws.cell(row, 1, f"── {section} ──")
            c.font = Font(bold=True, size=11, color=color)
            c.fill = cfill("F8F8F8")
            ws.merge_cells(f"A{row}:D{row}")
            ws.row_dimensions[row].height = 20; row += 1

        raw = (values or DEFAULTS).get(key, DEFAULTS.get(key))
        c_lbl = ws.cell(row, 1, label)
        c_lbl.font = Font(size=10)
        c_lbl.alignment = Alignment(horizontal='right', vertical='center')
        c_lbl.border = tborder()

        c_val = ws.cell(row, 2, raw)
        c_val.fill = cfill(AMBER); c_val.font = Font(bold=True, size=10)
        c_val.alignment = Alignment(horizontal='center', vertical='center')
        c_val.border = tborder()
        if vtype == 'money': c_val.number_format = '#,##0'
        elif vtype == 'pct': c_val.number_format = '0.0%'
        else:                c_val.number_format = '0'

        # Highlight Tax Law Year row specially
        if key == 'tax_law_year':
            c_lbl.fill = cfill("FFF3E0")
            c_val.fill = cfill("FF6F00")
            c_val.font = Font(bold=True, size=10, color=WHITE)

        c_note = ws.cell(row, 3, notes)
        c_note.font = Font(italic=True, size=9, color="666666")
        c_note.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 18; row += 1

    row += 1
    inst = ws.cell(row, 1,
        "HOW TO USE:\n"
        "1. Edit the amber cells above.  The orange 'Tax Law Year' cell controls which IRS/CMS\n"
        "   tables are used. If you enter a future year not yet in the database, the latest\n"
        "   available year is used automatically and flagged in the output title.\n"
        f"   Available years: {sorted(TAX_LAW_DB.keys())}\n"
        "2. Save this file.\n"
        "3. Run:  python3 build_retirement_planner.py  <this_file.xlsx>\n"
        "4. Results appear in the output tabs. This Inputs tab is NEVER overwritten.\n"
        "   The Tax Tables tab shows the exact law values used in each run."
    )
    inst.font = Font(size=10, color="333333")
    inst.alignment = Alignment(wrap_text=True, vertical='top')
    inst.fill = cfill("EEF7FF")
    ws.merge_cells(f"A{row}:D{row}")
    ws.row_dimensions[row].height = 110

# ═══════════════════════════════════════════════════════════════════════════════
# READ INPUTS FROM EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
LABEL_ALIASES = {
    # Maps old/alternate label text → canonical key, so old Inputs sheets still read correctly
    "Gross Withdrawal Rate":        "withdraw_rate",
    "Withdrawal Rate":              "withdraw_rate",
    "Annual Withdrawal Rate":       "withdraw_rate",
}

def read_inputs(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "Inputs" not in wb.sheetnames:
        raise ValueError(f"No 'Inputs' sheet found in {filepath}.")
    ws = wb["Inputs"]
    label_to_row = {cell.value.strip(): cell.row
                    for row in ws.iter_rows() for cell in row
                    if cell.column == 1 and isinstance(cell.value, str)}

    # Build a key→row lookup that handles both canonical and alias labels
    key_to_row = {}
    for label, row in label_to_row.items():
        # canonical labels
        for canon_label, key, *_ in INPUT_ROWS:
            if label == canon_label:
                key_to_row[key] = row
        # alias labels
        if label in LABEL_ALIASES:
            key_to_row[LABEL_ALIASES[label]] = row

    inputs = {}
    for label, key, vtype, _, _ in INPUT_ROWS:
        if key not in key_to_row:
            inputs[key] = DEFAULTS[key]; continue
        raw = ws.cell(key_to_row[key], 2).value
        if raw is None:
            inputs[key] = DEFAULTS[key]; continue
        if vtype == 'int':    inputs[key] = int(raw)
        elif vtype == 'money': inputs[key] = float(raw)
        elif vtype == 'pct':
            v = float(raw)
            inputs[key] = v if v <= 1.0 else v / 100.0
        else: inputs[key] = raw
    wb.close()
    return inputs

# ═══════════════════════════════════════════════════════════════════════════════
# TAX CALCULATION FUNCTIONS  (all accept tl=tax_law dict, base_yr, inf)
# ═══════════════════════════════════════════════════════════════════════════════
SS_CLAIMING_PCT = {62:0.70, 63:0.75, 64:0.80, 65:0.867, 66:0.933,
                   67:1.00, 68:1.08, 69:1.16, 70:1.24}

ACA_AGE_FACTORS = {
    21:1.000, 25:1.004, 30:1.127, 35:1.160, 40:1.278, 45:1.426,
    50:1.786, 55:2.004, 60:2.571, 64:3.000
}

RMD_FACTORS = {
    75:24.6, 76:23.7, 77:22.9, 78:22.0, 79:21.1, 80:20.2,
    81:19.4, 82:18.5, 83:17.7, 84:16.8, 85:16.0, 86:15.2,
    87:14.4, 88:13.7, 89:12.9, 90:12.2, 91:11.5, 92:10.8,
    93:10.1, 94:9.5,  95:8.9,  96:8.4,  97:7.8,  98:7.3
}

def inf_f(yr, base_yr, inf): return (1 + inf) ** (yr - base_yr)

def std_ded(yr, s1_age, s2_age, tl, base_yr, inf):
    f = inf_f(yr, base_yr, inf)
    base   = tl['std_ded_mfj'] * f
    senior = tl['senior_bonus'] * f * ((1 if s1_age>=65 else 0) + (1 if s2_age>=65 else 0))
    return base + senior

def fed_tax_ordinary(gross, yr, s1_age, s2_age, tl, base_yr, inf):
    f  = inf_f(yr, base_yr, inf)
    sd = std_ded(yr, s1_age, s2_age, tl, base_yr, inf)
    taxable = max(0.0, gross - sd); tax = 0.0
    for lo, hi, rate in tl['fed_brackets']:
        adj_lo = lo*f; adj_hi = hi*f if hi<1e11 else 1e12
        if taxable > adj_lo:
            tax += (min(taxable, adj_hi) - adj_lo) * rate
    return tax

def or_tax_ordinary(gross, yr, tl, base_yr, inf):
    f     = inf_f(yr, base_yr, inf)
    or_sd = tl['or_std_ded'] * f
    taxable = max(0.0, gross - or_sd); tax = 0.0
    for lo, hi, rate in tl['or_brackets']:
        adj_lo = lo*f; adj_hi = hi*f if hi<1e11 else 1e12
        if taxable > adj_lo:
            tax += (min(taxable, adj_hi) - adj_lo) * rate
    return tax

# 2026 HHS Poverty Guidelines — 48 contiguous states (source: HHS ASPE Jan 2026)
# Each additional person adds $5,680
FPL_2026 = {1: 15_960, 2: 21_640, 3: 27_320, 4: 33_000, 5: 38_680}
FPL_INCR_2026 = 5_680   # per additional person beyond 5

def fpl_for_size(household_size, yr, inf):
    """FPL for given household size, inflated from 2026 base at the given inflation rate.
    ACA coverage year Y uses FPL published for year Y (prior-year HHS guidelines, yr-1 base)."""
    size = max(1, min(household_size, 8))
    if size <= 5:
        base = FPL_2026[size]
    else:
        base = FPL_2026[5] + (size - 5) * FPL_INCR_2026
    # ACA uses prior-year FPL, so reference year is yr-1
    return base * (1 + inf) ** (yr - 1 - 2026)

def children_on_aca(yr, p):
    """Returns list of ages for children still on parents' ACA plan (birth_year != 0 and age < 26)."""
    kids = []
    for key in ('child1_birth', 'child2_birth', 'child3_birth'):
        birth = p.get(key, 0)
        if birth and birth > 0:
            age = yr - birth
            if age < 26:
                kids.append(age)
    return kids

def aca_household_size(yr, p):
    """Tax household size for ACA FPL calculation: 2 spouses + children still on plan."""
    return 2 + len(children_on_aca(yr, p))

def aca_subsidy_cliff(yr, s1_age, s2_age, p, inf):
    """
    MAGI at which ACA Premium Tax Credits phase out completely (400% FPL).
    Enhanced subsidies (ARPA/IRA) expired Jan 1 2026 — hard 400% cliff is back.
    Household size adjusts as children age off the plan at 26.
    Returns 0 when both spouses are on Medicare (ACA no longer relevant).
    """
    if s1_age >= 65 and s2_age >= 65:
        return 0.0   # both on Medicare — ACA subsidies N/A
    hh_size = aca_household_size(yr, p)
    return 4.0 * fpl_for_size(hh_size, yr, inf)

def bracket_ceiling(yr, s1_age, s2_age, target_rate, tl, base_yr, inf):
    f  = inf_f(yr, base_yr, inf)
    sd = std_ded(yr, s1_age, s2_age, tl, base_yr, inf)
    for lo, hi, rate in tl['fed_brackets']:
        if rate == target_rate:
            return hi * f + sd
    return tl['fed_brackets'][-2][1] * f + sd   # fallback: top of 35% bracket

def conv_bracket_taxable_top(yr, target_rate, tl, base_yr, inf):
    f = inf_f(yr, base_yr, inf)
    for lo, hi, rate in tl['fed_brackets']:
        if rate == target_rate:
            return hi * f
    return tl['fed_brackets'][-2][1] * f

def highest_bracket_fed(gross, yr, s1_age, s2_age, tl, base_yr, inf):
    f  = inf_f(yr, base_yr, inf)
    sd = std_ded(yr, s1_age, s2_age, tl, base_yr, inf)
    taxable = max(0.0, gross - 1 - sd); rate = tl['fed_brackets'][0][2]
    for lo, hi, r in tl['fed_brackets']:
        if taxable > lo*f: rate = r
    return rate

def highest_bracket_or(gross, yr, tl, base_yr, inf):
    f     = inf_f(yr, base_yr, inf)
    or_sd = tl['or_std_ded'] * f
    taxable = max(0.0, gross - 1 - or_sd); rate = tl['or_brackets'][0][2]
    for lo, hi, r in tl['or_brackets']:
        if taxable > lo*f: rate = r
    return rate

def ss_taxable_amount(ss_total, other_agi):
    prov = other_agi + 0.50 * ss_total
    if prov <= 32_000: return 0.0
    elif prov <= 44_000: return min(0.50*ss_total, 0.50*(prov-32_000))
    t1 = min(0.50*ss_total, 6_000)
    t2 = min(0.85*ss_total - t1, 0.85*(prov-44_000))
    return min(t1+t2, 0.85*ss_total)

def div_taxes(dividends, magi, yr, tl, base_yr, inf):
    f = inf_f(yr, base_yr, inf)
    ltcg_rate = 0.0
    for lo, hi, rate in tl['ltcg_brackets']:
        if magi > lo*f: ltcg_rate = rate
    fed  = dividends * ltcg_rate
    or_d = dividends * 0.099
    niit = dividends * 0.038 if magi > tl['niit_threshold'] else 0.0
    return fed, or_d, niit

def taxable_principal_taxes(amount, magi, yr, tl, base_yr, inf):
    f = inf_f(yr, base_yr, inf)
    ltcg_rate = 0.0
    for lo, hi, rate in tl['ltcg_brackets']:
        if magi > lo*f: ltcg_rate = rate
    fed  = amount * ltcg_rate
    or_c = amount * 0.099
    niit = amount * 0.038 if magi > tl['niit_threshold'] else 0.0
    return fed, or_c, niit

def irmaa_surcharge(magi, yr, s1_age, s2_age, tl, base_yr):
    persons = (1 if s1_age>=65 else 0) + (1 if s2_age>=65 else 0)
    if persons == 0: return 0.0
    f = (1.02) ** (yr - base_yr)
    # Find the first tier whose inflated threshold >= MAGI — that tier's surcharge applies.
    # Table is ordered: (upper_bound_of_tier, surcharge_for_tier).
    # e.g. (218K, 0) means MAGI ≤ 218K → $0; (274K, 1147.92) means 218K < MAGI ≤ 274K → $1,147.92
    surcharge_pp = tl['irmaa'][-1][1]   # default: top tier
    for thresh, annual_pp in tl['irmaa']:
        if magi <= thresh * f:
            surcharge_pp = annual_pp
            break
    return surcharge_pp * f * persons

def aca_age_factor(age):
    ages = sorted(ACA_AGE_FACTORS.keys())
    for i, a in enumerate(ages):
        if age <= a:
            if i == 0: return ACA_AGE_FACTORS[a]
            a0, a1 = ages[i-1], a
            return ACA_AGE_FACTORS[a0] + (ACA_AGE_FACTORS[a1]-ACA_AGE_FACTORS[a0])*(age-a0)/(a1-a0)
    return ACA_AGE_FACTORS[64]

def insurance_cost(yr, s1_age, s2_age, p):
    """Annual insurance cost: both spouses + any children still on parents' ACA plan."""
    yrs = yr - 2026
    aca_mo_base = p['aca_base_21'] * (1 + p['aca_inf']) ** yrs

    s1 = p['medicare_base_mo']*12*(1+p['med_inf'])**yrs if s1_age>=65 else \
         aca_mo_base * aca_age_factor(s1_age) * 12
    s2 = p['medicare_base_mo']*12*(1+p['med_inf'])**yrs if s2_age>=65 else \
         aca_mo_base * aca_age_factor(s2_age) * 12

    # Add ACA premiums for children still on the plan (under 26)
    # Under 21: ACA rates at age-21 (cannot charge more per ACA rules)
    # Ages 21-25: use adult age factors (nearly flat at 1.000-1.004)
    kids_cost = 0.0
    for child_age in children_on_aca(yr, p):
        effective_age = max(child_age, 21)   # ACA child rate = age-21 rate for under-21
        kids_cost += aca_mo_base * aca_age_factor(effective_age) * 12

    return s1 + s2 + kids_cost

def ss_annual(yr, p):
    pia1 = p['ss1_fra_mo']*12;  pia2 = p['ss2_fra_mo']*12
    s1_start = p['spouse1_birth'] + p['ss1_claim_age']
    s1_base  = pia1 * SS_CLAIMING_PCT[p['ss1_claim_age']]
    ss1 = s1_base * (1+p['ss_cola'])**(yr-s1_start) if yr>=s1_start else 0.0
    s2_start = p['spouse2_birth'] + p['ss2_claim_age']
    s2_base  = max(pia2 * SS_CLAIMING_PCT[p['ss2_claim_age']], pia1*0.50)
    ss2 = s2_base * (1+p['ss_cola'])**(yr-s2_start) if yr>=s2_start else 0.0
    return ss1, ss2, ss1+ss2

def rmd_amount(trad_prev, owner_age, tl):
    rmd_age = tl.get('rmd_start_age', 75)
    if owner_age < rmd_age: return 0.0
    return trad_prev / RMD_FACTORS.get(min(owner_age, 98), 7.3)

def get_phase_info(i, p):
    """Return (phase_name, phase_rate, is_phase_start) for simulation year index i."""
    gogo_yrs   = p.get('gogo_years',   0) or p['retire_years'] // 3
    slowgo_yrs = p.get('slowgo_years', 0) or p['retire_years'] // 3
    # nogo gets the remainder
    gogo_rate   = p.get('gogo_rate',   0) or p.get('withdraw_rate', 0.05)
    slowgo_rate = p.get('slowgo_rate', 0) or p.get('withdraw_rate', 0.04)
    nogo_rate   = p.get('nogo_rate',   0) or p.get('withdraw_rate', 0.03)

    if i < gogo_yrs:
        return 'Go-Go',   gogo_rate,   (i == 0)
    elif i < gogo_yrs + slowgo_yrs:
        return 'Slow-Go', slowgo_rate, (i == gogo_yrs)
    else:
        return 'No-Go',   nogo_rate,   (i == gogo_yrs + slowgo_yrs)


def _trial_after_tax_cash(gross_target_trial, ann_div, ss_total, s1_age, s2_age, yr,
                           trad_bal, roth_bal, taxable_bal, p, tl, base_yr, inf):
    """Compute after_tax_cash for a trial gross_target — used by binary search."""
    ss_tax_est = ss_total * 0.85
    cap_ord    = bracket_ceiling(yr, s1_age, s2_age, p['conv_bracket'], tl, base_yr, inf)
    max_trad   = min(max(0.0, cap_ord - ss_tax_est), trad_bal)
    trad_needed = max(0.0, gross_target_trial - ann_div)
    trad_wd    = min(trad_needed, max_trad)
    remaining  = max(0.0, trad_needed - trad_wd)
    taxable_principal_wd = min(remaining, max(0.0, taxable_bal))
    roth_wd    = min(max(0.0, remaining - taxable_principal_wd), roth_bal)
    ss_tax_amt = ss_taxable_amount(ss_total, trad_wd + ann_div)
    sd         = std_ded(yr, s1_age, s2_age, tl, base_yr, inf)
    used_taxable = max(0.0, trad_wd + ss_tax_amt - sd)
    if p['conv_bracket'] > 0:
        conv_top  = conv_bracket_taxable_top(yr, p['conv_bracket'], tl, base_yr, inf)
        roth_conv = min(max(0.0, conv_top - used_taxable), max(0.0, trad_bal - trad_wd))
    else:
        roth_conv = 0.0
    ordinary_gross = trad_wd + roth_conv + ss_tax_amt
    f_tax_ord  = fed_tax_ordinary(ordinary_gross, yr, s1_age, s2_age, tl, base_yr, inf)
    or_tax_ord = or_tax_ordinary(ordinary_gross, yr, tl, base_yr, inf)
    magi       = trad_wd + roth_conv + ss_tax_amt + ann_div + taxable_principal_wd
    div_fed, div_or, div_niit = div_taxes(ann_div, magi, yr, tl, base_yr, inf)
    tx_fed, tx_or, tx_niit   = taxable_principal_taxes(taxable_principal_wd, magi, yr, tl, base_yr, inf)
    # Conversion taxes paid from conversion — only spending taxes reduce cash
    f_tax_spend  = fed_tax_ordinary(trad_wd + ss_tax_amt, yr, s1_age, s2_age, tl, base_yr, inf)
    or_tax_spend = or_tax_ordinary(trad_wd + ss_tax_amt, yr, tl, base_yr, inf)
    trad_net              = trad_wd - f_tax_spend - or_tax_spend
    div_net               = ann_div - (div_fed + div_or + div_niit)
    taxable_principal_net = taxable_principal_wd - (tx_fed + tx_or + tx_niit)
    return trad_net + roth_wd + div_net + taxable_principal_net + ss_total


# ═══════════════════════════════════════════════════════════════════════════════
# CORE SIMULATION
# ═══════════════════════════════════════════════════════════════════════════════
def run_simulation(p, tl, base_yr):
    rows = []
    trad_bal    = p['trad_bal']
    roth_bal    = p['roth_bal']
    taxable_bal = p['taxable_bal']
    trad_bal_prev = p['trad_bal']
    inf = p['inflation']
    prev_gross_target = None
    prev_atc_target   = None
    prev_phase_rate   = None
    transition_remaining  = 0     # years left in current taper
    transition_total_yrs  = 0     # total length of current taper (for interpolation)
    transition_start_gross  = None  # gross at start of taper
    transition_target_gross = None  # gross at end of taper
    post_taper = False            # True after a taper completes — phase boundaries inflate, no re-anchor

    for i in range(p['retire_years']):
        yr     = p['retire_year'] + i
        s1_age = p['spouse1_age'] + (yr - 2026)
        s2_age = p['spouse2_age'] + (yr - 2026)

        ss1, ss2, ss_total = ss_annual(yr, p)
        ann_div    = taxable_bal * p['div_yield']
        rmd        = rmd_amount(trad_bal_prev, s1_age, tl)
        total_port = trad_bal + roth_bal + taxable_bal
        # Year 1: gross target from portfolio × withdrawal rate (or binary-search to hit
        #         an after-tax cash target if after_tax_rate is set).
        # Years 2+: inflate prior year's gross target (inflation-adjusted spending).
        phase_name, phase_rate, is_phase_start = get_phase_info(i, p)
        after_tax_rate = p.get('after_tax_rate', 0.0)

        trans_yrs = p.get('transition_years', 0)

        if after_tax_rate > 0:
            # Binary search every year.
            # Year 1: ATC target = portfolio × phase_rate.
            # Phase transition (i > 0): scale prior ATC target by rate ratio for smooth drop.
            # Within phase: inflate from prior year.
            if i == 0:
                atc_target = total_port * phase_rate
            elif is_phase_start:
                atc_target = prev_atc_target * (phase_rate / prev_phase_rate)
            else:
                atc_target = prev_atc_target * (1 + inf)
            lo, hi = 0.0, total_port
            for _ in range(25):
                mid = (lo + hi) / 2
                atc = _trial_after_tax_cash(mid, ann_div, ss_total, s1_age, s2_age, yr,
                                            trad_bal, roth_bal, taxable_bal, p, tl, base_yr, inf)
                if atc < atc_target: lo = mid
                else:                hi = mid
            gross_target = (lo + hi) / 2
            active_rate = phase_rate

        elif i == 0:
            # First year: anchor to portfolio × rate
            gross_target = total_port * phase_rate
            active_rate  = phase_rate
            transition_remaining  = 0
            transition_total_yrs  = 0

        elif transition_remaining > 0:
            # Mid-taper: continue gliding — takes priority over phase boundaries
            # so Slow-Go→No-Go doesn't restart a second taper mid-glide.
            t = transition_total_yrs - transition_remaining + 1
            gross_target = transition_start_gross + (transition_target_gross - transition_start_gross) * t / transition_total_yrs
            active_rate  = phase_rate
            transition_remaining -= 1
            if transition_remaining == 0:
                post_taper = True  # taper just finished; subsequent phases inflate, no re-anchor

        elif is_phase_start and trans_yrs > 0 and phase_name == 'Slow-Go':
            # Taper only fires at Go-Go → Slow-Go transition.
            # Slow-Go → No-Go re-anchors normally (handled by the elif below).
            # If trans_yrs covers the full Slow-Go + No-Go span (i.e. user wants one
            # continuous glide), use the No-Go rate as the terminal target so spending
            # lands at the correct long-run level at the end of retirement.
            gogo_yrs_p   = p.get('gogo_years',   0) or p['retire_years'] // 3
            slowgo_yrs_p = p.get('slowgo_years',  0) or p['retire_years'] // 3
            nogo_yrs_p   = p['retire_years'] - gogo_yrs_p - slowgo_yrs_p
            nogo_rate_p  = p.get('nogo_rate', 0) or p.get('withdraw_rate', 0.03)

            if phase_name == 'Slow-Go' and trans_yrs >= slowgo_yrs_p + nogo_yrs_p:
                # Extended taper: user explicitly set taper to cover ALL remaining years —
                # glide all the way to No-Go level in one continuous glide.
                terminal_rate        = nogo_rate_p
                transition_total_yrs = slowgo_yrs_p + nogo_yrs_p
            else:
                terminal_rate        = phase_rate
                transition_total_yrs = trans_yrs

            proportional  = prev_gross_target * (terminal_rate / prev_phase_rate)
            portfolio_cap = total_port * terminal_rate
            transition_target_gross = min(proportional, portfolio_cap)
            transition_start_gross  = prev_gross_target
            transition_remaining    = transition_total_yrs
            t = 1
            gross_target = transition_start_gross + (transition_target_gross - transition_start_gross) * t / transition_total_yrs
            active_rate  = phase_rate
            transition_remaining -= 1

        elif is_phase_start and not post_taper:
            # Instant step-down (trans_yrs == 0, or no taper active)
            proportional  = prev_gross_target * (phase_rate / prev_phase_rate)
            portfolio_cap = total_port * phase_rate
            gross_target  = min(proportional, portfolio_cap)
            active_rate   = phase_rate
            transition_remaining = 0

        else:
            # Within a phase, or post-taper phase boundary: inflate from prior year
            gross_target = prev_gross_target * (1 + inf)
            active_rate  = phase_rate

        # Step 1: Trad (up to conv_bracket ceiling)
        ss_tax_est = ss_total * 0.85
        cap_ord    = bracket_ceiling(yr, s1_age, s2_age, p['conv_bracket'], tl, base_yr, inf)
        max_trad   = min(max(0.0, cap_ord - ss_tax_est), trad_bal)
        trad_needed = max(0.0, gross_target - ann_div)
        trad_wd    = min(trad_needed, max_trad)
        trad_wd    = max(trad_wd, min(rmd, trad_bal))

        # Step 2: Taxable principal
        remaining  = max(0.0, trad_needed - trad_wd)
        taxable_principal_wd = min(remaining, max(0.0, taxable_bal))

        # Step 3: Roth
        roth_wd    = min(max(0.0, remaining - taxable_principal_wd), roth_bal)

        # Roth conversion: fill conv_bracket headroom (skip if conv_bracket=0)
        ss_tax_amt  = ss_taxable_amount(ss_total, trad_wd + ann_div)
        ordinary_wd = trad_wd + ss_tax_amt
        sd          = std_ded(yr, s1_age, s2_age, tl, base_yr, inf)
        used_taxable = max(0.0, ordinary_wd - sd)
        if p['conv_bracket'] > 0:
            conv_top  = conv_bracket_taxable_top(yr, p['conv_bracket'], tl, base_yr, inf)
            roth_conv = min(max(0.0, conv_top - used_taxable), max(0.0, trad_bal - trad_wd))
        else:
            roth_conv = 0.0

        # Taxes
        ordinary_gross = trad_wd + roth_conv + ss_tax_amt
        f_tax_ord  = fed_tax_ordinary(ordinary_gross, yr, s1_age, s2_age, tl, base_yr, inf)
        or_tax_ord = or_tax_ordinary(ordinary_gross, yr, tl, base_yr, inf)
        magi       = trad_wd + roth_conv + ss_tax_amt + ann_div + taxable_principal_wd
        div_fed, div_or, div_niit = div_taxes(ann_div, magi, yr, tl, base_yr, inf)
        tx_fed, tx_or, tx_niit   = taxable_principal_taxes(taxable_principal_wd, magi, yr, tl, base_yr, inf)
        div_tax_total = div_fed + div_or + div_niit
        tx_tax_total  = tx_fed  + tx_or  + tx_niit
        total_tax     = f_tax_ord + or_tax_ord + div_tax_total + tx_tax_total

        # Split ordinary taxes: spending portion (trad_wd + SS) vs conversion portion.
        # Conversion taxes are paid from the conversion amount itself — they do NOT
        # reduce after-tax spendable cash; instead they reduce the net amount going into Roth.
        f_tax_spend  = fed_tax_ordinary(trad_wd + ss_tax_amt, yr, s1_age, s2_age, tl, base_yr, inf)
        or_tax_spend = or_tax_ordinary(trad_wd + ss_tax_amt, yr, tl, base_yr, inf)
        conv_tax     = (f_tax_ord + or_tax_ord) - (f_tax_spend + or_tax_spend)

        trad_net              = trad_wd - f_tax_spend - or_tax_spend   # only spending taxes here
        div_net               = ann_div - div_tax_total
        taxable_principal_net = taxable_principal_wd - tx_tax_total
        after_tax_cash        = trad_net + roth_wd + div_net + taxable_principal_net + ss_total
        ins_base      = insurance_cost(yr, s1_age, s2_age, p)
        irmaa         = irmaa_surcharge(magi, yr, s1_age, s2_age, tl, base_yr)
        ins           = ins_base + irmaa          # total insurance incl. IRMAA surcharge
        after_tax_net = after_tax_cash - ins
        fed_taxable_income = max(0.0, ordinary_gross - sd)
        ss_pct_taxed = (ss_tax_amt / ss_total) if ss_total > 0 else 0.0
        total_cash_in = trad_wd + roth_wd + ann_div + taxable_principal_wd + ss_total
        # eff1: ordinary income taxes only ÷ ordinary gross income
        #       (fed+OR on Trad WD/Roth conv/SS — what bracket management watches)
        #       Never exceeds ~50%; keeping numerator and denominator consistent
        eff1 = (f_tax_ord + or_tax_ord) / ordinary_gross if ordinary_gross > 0 else 0.0
        # eff2: ALL taxes ÷ ALL cash received (lifestyle rate)
        eff2 = total_tax / total_cash_in if total_cash_in > 0 else 0.0

        end_trad    = max(0.0, trad_bal - trad_wd - roth_conv) * (1 + p['growth_rate'])
        end_roth    = max(0.0, roth_bal - roth_wd + roth_conv - conv_tax) * (1 + p['growth_rate'])
        end_taxable = max(0.0, taxable_bal - taxable_principal_wd) * (1 + p['price_return'])
        end_port    = end_trad + end_roth + end_taxable

        rows.append({
            'year':yr, 's1_age':s1_age, 's2_age':s2_age, 'phase':phase_name,
            'trad_open':trad_bal, 'roth_open':roth_bal, 'taxable_open':taxable_bal,
            'total_port':total_port, 'end_port':end_port, 'gross_target':gross_target,
            'trad_wd':trad_wd, 'taxable_principal_wd':taxable_principal_wd,
            'roth_wd':roth_wd, 'div':ann_div, 'roth_conv':roth_conv,
            'ss1':ss1, 'ss2':ss2, 'ss_total':ss_total, 'ss_tax_amt':ss_tax_amt,
            'ordinary_gross':ordinary_gross, 'magi':magi,
            'fed_taxable_income':fed_taxable_income, 'ss_pct_taxed':ss_pct_taxed,
            'f_tax_ord':f_tax_ord, 'or_tax_ord':or_tax_ord, 'conv_tax':conv_tax,
            'div_fed':div_fed, 'div_or':div_or, 'div_niit':div_niit,
            'div_tax_total':div_tax_total,
            'tx_fed':tx_fed, 'tx_or':tx_or, 'tx_niit':tx_niit,
            'tx_tax_total':tx_tax_total, 'total_tax':total_tax,
            'hb_fed':highest_bracket_fed(ordinary_gross, yr, s1_age, s2_age, tl, base_yr, inf),
            'hb_or': highest_bracket_or(ordinary_gross, yr, tl, base_yr, inf),
            'irmaa':irmaa, 'eff1':eff1, 'eff2':eff2,
            'after_tax_cash':after_tax_cash, 'ins':ins,
            'after_tax_net':after_tax_net, 'rmd':rmd,
            'active_rate':  active_rate,
            'actual_wd_rate': gross_target / total_port if total_port > 0 else 0.0,
            'aca_hh_size': aca_household_size(yr, p) if not (s1_age>=65 and s2_age>=65) else 0,
            'aca_cliff':   aca_subsidy_cliff(yr, s1_age, s2_age, p, inf),
        })

        prev_gross_target = gross_target
        prev_atc_target   = atc_target if after_tax_rate > 0 else None
        prev_phase_rate   = phase_rate
        trad_bal_prev = trad_bal
        trad_bal    = max(0.0, trad_bal - trad_wd - roth_conv) * (1 + p['growth_rate'])
        roth_bal    = max(0.0, roth_bal - roth_wd  + roth_conv - conv_tax) * (1 + p['growth_rate'])
        taxable_bal = max(0.0, taxable_bal - taxable_principal_wd) * (1 + p['price_return'])

    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# TAX TABLES SHEET  — shows exact law values used in this run
# ═══════════════════════════════════════════════════════════════════════════════
def write_tax_tables(wb, tl, base_yr, actual_yr, fallback_used, p):
    if "Tax Tables" in wb.sheetnames: del wb["Tax Tables"]
    ws = wb.create_sheet("Tax Tables")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ORANGE_H

    fallback_note = f"  ⚠  Requested {p['tax_law_year']} — not in database, using {actual_yr} (latest available)" if fallback_used else ""
    t = ws.cell(1, 1, f"TAX TABLES USED IN THIS RUN — Base Year: {actual_yr}{fallback_note}")
    t.font = Font(bold=True, size=13, color=ORANGE_H if not fallback_used else RED_H)
    ws.merge_cells("A1:E1")

    ws.cell(2, 1, f"All thresholds are inflated at {p['inflation']*100:.1f}%/yr from {actual_yr} forward. "
                  f"Rates are statutory and do not change. NIIT threshold ${tl['niit_threshold']:,} and SS thresholds "
                  f"($32K/$44K) are not inflation-adjusted per statute.").font = Font(italic=True, size=9, color="555555")
    ws.merge_cells("A2:E2")

    r = 4

    def section_hdr(row, title, color):
        c = ws.cell(row, 1, title)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = hfill(color)
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = 20
        return row + 1

    def col_hdrs(row, headers, color):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.fill = hfill(color); c.font = Font(bold=True, color=WHITE, size=9)
            c.alignment = Alignment(horizontal='center')
            c.border = tborder()
        return row + 1

    def data_row(row, values, fmts, shade=False):
        for ci, (val, fmt) in enumerate(zip(values, fmts), 1):
            c = ws.cell(row, ci, val)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = tborder()
            if fmt == '$': c.number_format = '#,##0'
            if fmt == '%': c.number_format = '0.00%'
            if shade: c.fill = cfill("F5F5F5")
        return row + 1

    inf = p['inflation']

    # ── Federal Brackets ──
    r = section_hdr(r, "FEDERAL INCOME TAX — MFJ (TCJA Extended)", BLUE_H)
    r = col_hdrs(r, [f"Taxable Income (from) — {actual_yr}",
                     f"Taxable Income (to) — {actual_yr}",
                     f"Taxable Income (to) — {p['retire_year']}",
                     f"Taxable Income (to) — {p['retire_year']+17}",
                     "Rate"], BLUE_H)
    for i, (lo, hi, rate) in enumerate(tl['fed_brackets']):
        f0 = inf_f(p['retire_year'], base_yr, inf)
        f1 = inf_f(p['retire_year']+17, base_yr, inf)
        hi_disp  = hi if hi < 1e11 else "No limit"
        hi_r0    = f"${hi*f0:,.0f}" if hi < 1e11 else "No limit"
        hi_r1    = f"${hi*f1:,.0f}" if hi < 1e11 else "No limit"
        vals = [f"${lo:,}", f"${hi:,}" if hi<1e11 else "No limit", hi_r0, hi_r1, rate]
        fmts = ['', '', '', '', '%']
        rr = ws.cell(r, 1, f"${lo:,}");  rr.border = tborder(); rr.alignment = Alignment(horizontal='right')
        ws.cell(r, 2, f"${hi:,}" if hi<1e11 else "No limit").border = tborder()
        ws.cell(r, 3, hi_r0).border = tborder()
        ws.cell(r, 4, hi_r1).border = tborder()
        c_rate = ws.cell(r, 5, rate); c_rate.number_format = '0%'; c_rate.border = tborder()
        c_rate.font = Font(bold=True)
        if i % 2 == 0:
            for ci in range(1,6): ws.cell(r, ci).fill = cfill("F5F5F5")
        r += 1

    r += 1

    # ── Standard Deduction ──
    r = section_hdr(r, "STANDARD DEDUCTION (MFJ, inflation-adjusted)", TEAL_H)
    r = col_hdrs(r, [f"Item", f"Amount ({actual_yr})", f"Amount ({p['retire_year']})",
                     f"Amount ({p['retire_year']+17})", "Notes"], TEAL_H)
    sd_items = [
        ("Base (MFJ)",         tl['std_ded_mfj'],  "Both spouses"),
        ("Senior Bonus (65+)", tl['senior_bonus'],  "Per eligible spouse, added at age 65"),
    ]
    for i, (label, val, note) in enumerate(sd_items):
        f0 = inf_f(p['retire_year'], base_yr, inf)
        f1 = inf_f(p['retire_year']+17, base_yr, inf)
        ws.cell(r, 1, label).border = tborder()
        ws.cell(r, 2, val).number_format='#,##0'; ws.cell(r,2).border=tborder()
        ws.cell(r, 3, val*f0).number_format='#,##0'; ws.cell(r,3).border=tborder()
        ws.cell(r, 4, val*f1).number_format='#,##0'; ws.cell(r,4).border=tborder()
        ws.cell(r, 5, note).border=tborder(); ws.cell(r,5).font=Font(italic=True,size=9,color="666666")
        if i%2==0:
            for ci in range(1,6): ws.cell(r,ci).fill=cfill("F5F5F5")
        r += 1

    r += 1

    # ── LTCG Brackets ──
    r = section_hdr(r, "LONG-TERM CAPITAL GAINS / QUALIFIED DIVIDENDS (MFJ)", GREEN_H)
    r = col_hdrs(r, [f"Taxable Income (from) — {actual_yr}",
                     f"Taxable Income (to) — {actual_yr}",
                     f"To — {p['retire_year']}",
                     f"To — {p['retire_year']+17}", "LTCG Rate"], GREEN_H)
    for i, (lo, hi, rate) in enumerate(tl['ltcg_brackets']):
        f0 = inf_f(p['retire_year'], base_yr, inf)
        f1 = inf_f(p['retire_year']+17, base_yr, inf)
        ws.cell(r,1,f"${lo:,}").border=tborder()
        ws.cell(r,2,f"${hi:,}" if hi<1e11 else "No limit").border=tborder()
        ws.cell(r,3,f"${hi*f0:,.0f}" if hi<1e11 else "No limit").border=tborder()
        ws.cell(r,4,f"${hi*f1:,.0f}" if hi<1e11 else "No limit").border=tborder()
        c=ws.cell(r,5,rate); c.number_format='0%'; c.font=Font(bold=True); c.border=tborder()
        if i%2==0:
            for ci in range(1,6): ws.cell(r,ci).fill=cfill("F5F5F5")
        r += 1

    r += 1

    # ── Oregon Brackets ──
    r = section_hdr(r, "OREGON STATE INCOME TAX (MFJ)", RED_H)
    r = col_hdrs(r, [f"OR Taxable Income (from) — {actual_yr}",
                     f"OR Taxable Income (to) — {actual_yr}",
                     f"To — {p['retire_year']}",
                     f"To — {p['retire_year']+17}", "OR Rate"], RED_H)
    for i, (lo, hi, rate) in enumerate(tl['or_brackets']):
        f0 = inf_f(p['retire_year'], base_yr, inf)
        f1 = inf_f(p['retire_year']+17, base_yr, inf)
        ws.cell(r,1,f"${lo:,}").border=tborder()
        ws.cell(r,2,f"${hi:,}" if hi<1e11 else "No limit").border=tborder()
        ws.cell(r,3,f"${hi*f0:,.0f}" if hi<1e11 else "No limit").border=tborder()
        ws.cell(r,4,f"${hi*f1:,.0f}" if hi<1e11 else "No limit").border=tborder()
        c=ws.cell(r,5,rate); c.number_format='0.00%'; c.font=Font(bold=True); c.border=tborder()
        if i%2==0:
            for ci in range(1,6): ws.cell(r,ci).fill=cfill("F5F5F5")
        r += 1
    r += 1
    ws.cell(r,1,"Oregon Standard Deduction").border=tborder()
    ws.cell(r,2,tl['or_std_ded']).number_format='#,##0'; ws.cell(r,2).border=tborder()
    ws.cell(r,3,tl['or_std_ded']*inf_f(p['retire_year'],base_yr,inf)).number_format='#,##0'; ws.cell(r,3).border=tborder()
    ws.cell(r,4,tl['or_std_ded']*inf_f(p['retire_year']+17,base_yr,inf)).number_format='#,##0'; ws.cell(r,4).border=tborder()
    ws.cell(r,5,"Inflation-adjusted").border=tborder(); r+=1

    r += 1

    # ── IRMAA ──
    r = section_hdr(r, "IRMAA — Medicare Part B/D Surcharges (MFJ, 2-year lookback)", PURPLE_H)
    r = col_hdrs(r, [f"MAGI Threshold — {actual_yr}",
                     f"MAGI Threshold — {p['retire_year']}",
                     f"MAGI Threshold — {p['retire_year']+17}",
                     f"Annual Surcharge/person — {actual_yr}", "Notes"], PURPLE_H)
    for i, (thresh, surcharge_pp) in enumerate(tl['irmaa']):
        f0 = (1.02)**(p['retire_year']-base_yr)
        f1 = (1.02)**(p['retire_year']+17-base_yr)
        thresh_disp = f"${thresh:,.0f}" if thresh < 1e11 else "Over $750K"
        t0 = f"${thresh*f0:,.0f}" if thresh < 1e11 else "Over limit"
        t1 = f"${thresh*f1:,.0f}" if thresh < 1e11 else "Over limit"
        ws.cell(r,1,thresh_disp).border=tborder()
        ws.cell(r,2,t0).border=tborder()
        ws.cell(r,3,t1).border=tborder()
        ws.cell(r,4,surcharge_pp).number_format='#,##0'; ws.cell(r,4).border=tborder()
        ws.cell(r,5,"Thresholds inflate 2%/yr per CMS").border=tborder()
        ws.cell(r,5).font=Font(italic=True,size=9,color="666666")
        if i%2==0:
            for ci in range(1,6): ws.cell(r,ci).fill=cfill("F5F5F5")
        r += 1

    r += 1

    # ── Fixed constants ──
    r = section_hdr(r, "FIXED STATUTORY AMOUNTS (not inflation-adjusted)", GREY_H)
    fixed = [
        ("NIIT Rate", "3.8%", f"Applies when MAGI > ${tl['niit_threshold']:,} (MFJ, not indexed)"),
        ("SS Taxability Threshold 1", "$32,000 (MFJ)", "Below: 0% of SS taxable"),
        ("SS Taxability Threshold 2", "$44,000 (MFJ)", "Above: up to 85% of SS taxable"),
        ("RMD Start Age", str(tl.get('rmd_start_age',75)), "SECURE 2.0 — born 1960+"),
        (f"QCD Annual Limit ({actual_yr})", f"${tl['qcd_limit']:,}/person", "Age 70½+, indexed to inflation"),
    ]
    for i, (label, val, note) in enumerate(fixed):
        ws.cell(r,1,label).border=tborder(); ws.cell(r,1).font=Font(bold=True,size=9)
        ws.cell(r,2,val).border=tborder(); ws.cell(r,2).alignment=Alignment(horizontal='center')
        ws.cell(r,3,note).border=tborder(); ws.cell(r,3).font=Font(italic=True,size=9,color="666666")
        ws.merge_cells(f"C{r}:E{r}")
        if i%2==0:
            for ci in range(1,6): ws.cell(r,ci).fill=cfill("F5F5F5")
        r += 1

    for ci, w in enumerate([30, 22, 22, 22, 36], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# REMAINING OUTPUT SHEETS  (same as before, pass tl/base_yr where needed)
# ═══════════════════════════════════════════════════════════════════════════════
def write_assumptions_output(wb, rows, p, tl, base_yr, actual_yr, fallback_used):
    if "Assumptions" in wb.sheetnames: del wb["Assumptions"]
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREY_H

    def row_out(r, label, value, note="", bold_val=False):
        c_lbl = ws.cell(r, 1, label)
        c_lbl.font = Font(bold=True, size=10)
        c_lbl.alignment = Alignment(horizontal='right', vertical='center')
        c_lbl.border = tborder()
        c_val = ws.cell(r, 2, value)
        c_val.alignment = Alignment(horizontal='left', vertical='center')
        c_val.border = tborder()
        if bold_val: c_val.font = Font(bold=True, size=10)
        if note:
            ws.cell(r, 3, note).font = Font(italic=True, color="666666", size=9)
        return c_val

    fallback_note = f" (⚠ requested {p['tax_law_year']}, using {actual_yr} — not yet in database)" if fallback_used else ""
    ws.cell(1, 1, "ASSUMPTIONS USED IN THIS RUN").font = Font(bold=True, size=13, color=BLUE_H)
    ws.merge_cells("A1:C1")
    rate_desc = f"Go-Go {p.get('gogo_rate',0)*100:.0f}% / Slow-Go {p.get('slowgo_rate',0)*100:.0f}% / No-Go {p.get('nogo_rate',0)*100:.0f}%" \
                if p.get('gogo_rate', 0) > 0 else f"{p.get('withdraw_rate',0)*100:.0f}%"
    ws.cell(2, 1, f"Generated: {datetime.date.today()} | Withdrawal: {rate_desc} | Conv Bracket: {p['conv_bracket']*100:.0f}% | Tax Law: {actual_yr}{fallback_note}")
    ws.cell(2, 1).font = Font(italic=True, size=9, color="888888" if not fallback_used else RED_H)
    ws.merge_cells("A2:C2")

    r = 4

    def section(row, title, color):
        c = ws.cell(row, 1, f"── {title} ──")
        c.font = Font(bold=True, color=color, size=11)
        ws.merge_cells(f"A{row}:C{row}")
        return row + 1

    r = section(r, "PARAMETERS READ FROM INPUTS SHEET", BLUE_H)
    for label, key, vtype, notes, _ in INPUT_ROWS:
        val = p.get(key)
        if vtype == 'pct':     disp = f"{val*100:.1f}%"
        elif vtype == 'money': disp = f"${val:,.0f}"
        else:                  disp = str(val)
        row_out(r, label, disp, notes)
        r += 1

    r += 1; r = section(r, "35-YEAR SUMMARY TOTALS", BLUE_H)
    sums = [
        ("Total Gross Withdrawal",   sum(x['gross_target'] for x in rows),         '$'),
        ("  From Tax-Deferred",       sum(x['trad_wd'] for x in rows),              '$'),
        ("  From Taxable Principal",  sum(x['taxable_principal_wd'] for x in rows), '$'),
        ("  From Roth",               sum(x['roth_wd'] for x in rows),              '$'),
        ("  Taxable Dividends",       sum(x['div'] for x in rows),                  '$'),
        ("Total Social Security",    sum(x['ss_total'] for x in rows),              '$'),
        ("Total Roth Conversions",   sum(x['roth_conv'] for x in rows),             '$'),
        ("Total Federal Tax",        sum(x['f_tax_ord'] for x in rows),             '$'),
        ("Total Oregon Tax",         sum(x['or_tax_ord'] for x in rows),            '$'),
        ("Total Dividend Tax",       sum(x['div_tax_total'] for x in rows),         '$'),
        ("Total Taxable LTCG Tax",   sum(x['tx_tax_total'] for x in rows),          '$'),
        ("Total All Taxes",          sum(x['total_tax'] for x in rows),             '$'),
        ("Total Insurance",          sum(x['ins'] for x in rows),                   '$'),
        ("Total Net Spendable",      sum(x['after_tax_net'] for x in rows),         '$'),
    ]
    for label, val, fmt in sums:
        cv = row_out(r, label, val, bold_val=label.startswith("Total"))
        cv.number_format = '#,##0'; r += 1

    total_cash = sum(x['trad_wd']+x['roth_wd']+x['div']+x['taxable_principal_wd']+x['ss_total'] for x in rows)
    cv = row_out(r, "Average Effective Tax Rate", sum(x['total_tax'] for x in rows)/total_cash if total_cash else 0)
    cv.number_format = '0.0%'; r += 1

    r += 1; r = section(r, "END BALANCES", GREY_H)
    last = rows[-1]
    for label, val in [("Tax-Deferred", max(0.0, last['trad_open'] - last['trad_wd'] - last['roth_conv']) * (1+p['growth_rate'])),
                        ("Roth",          max(0.0, last['roth_open'] - last['roth_wd'] + last['roth_conv'] - last['conv_tax']) * (1+p['growth_rate'])),
                        ("Taxable",       max(0.0, last['taxable_open'] - last['taxable_principal_wd']) * (1+p['price_return']))]:
        cv = row_out(r, label, val); cv.number_format = '#,##0'; r += 1

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 55


def write_annual_summary(wb, rows, p, actual_yr, fallback_used):
    if "Annual Summary" in wb.sheetnames: del wb["Annual Summary"]
    ws = wb.create_sheet("Annual Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D5"
    ws.sheet_properties.tabColor = GREEN_H

    COLS = [
        ("Year",               'year',                'int',  "YEAR",             6),
        ("S1\nAge",            's1_age',               'int',  "YEAR",             6),
        ("S2\nAge",            's2_age',               'int',  "YEAR",             6),
        ("Phase",              'phase',                'str',  "YEAR",             9),
        ("Actual\nW/D Rate",    'actual_wd_rate',        'pct',  "YEAR",             8),
        ("Trad IRA\nBalance",  'trad_open',            '$',    "YEAR-END BALANCES",14),
        ("Roth\nBalance",      'roth_open',            '$',    "YEAR-END BALANCES",13),
        ("Taxable\nBalance",   'taxable_open',         '$',    "YEAR-END BALANCES",14),
        ("Ending\nPortfolio\nBalance", 'end_port',     '$',    "YEAR-END BALANCES",16),
        ("Trad\nWithdrawal",   'trad_wd',              '$',    "WITHDRAWALS",      13),
        ("Taxable\nPrincipal", 'taxable_principal_wd', '$',    "WITHDRAWALS",      12),
        ("Roth\nWithdrawal",   'roth_wd',              '$',    "WITHDRAWALS",      12),
        ("Taxable\nDividends", 'div',                  '$',    "WITHDRAWALS",      12),
        ("Roth\nConversion",   'roth_conv',            '$',    "WITHDRAWALS",      12),
        ("SS\nSpouse 1",       'ss1',                  '$',    "WITHDRAWALS",      11),
        ("SS\nSpouse 2",       'ss2',                  '$',    "WITHDRAWALS",      11),
        ("Total Portfolio\nWithdrawal", 'gross_target', '$',    "WITHDRAWALS",      14),
        ("Hi Fed\nBracket",    'hb_fed',               'pct',  "ORDINARY TAX",     9),
        ("Federal\nTax (ord)", 'f_tax_ord',            '$',    "ORDINARY TAX",     13),
        ("Hi OR\nBracket",     'hb_or',                'pct',  "ORDINARY TAX",     9),
        ("Oregon\nTax (ord)",  'or_tax_ord',           '$',    "ORDINARY TAX",     13),
        ("Federal\nTaxable Inc",'fed_taxable_income',  '$',    "ORDINARY TAX",     13),
        ("% SS\nTaxed",        'ss_pct_taxed',         'pct',  "ORDINARY TAX",     9),
        ("Div Tax\nFed LTCG",  'div_fed',              '$',    "DIVIDEND TAX",     12),
        ("Div Tax\nOR 9.9%",   'div_or',               '$',    "DIVIDEND TAX",     12),
        ("Div Tax\nNIIT 3.8%", 'div_niit',             '$',    "DIVIDEND TAX",     11),
        ("Taxable LTCG\nFed",  'tx_fed',               '$',    "TAXABLE LTCG TAX", 12),
        ("Taxable LTCG\nOR",   'tx_or',                '$',    "TAXABLE LTCG TAX", 12),
        ("Taxable LTCG\nNIIT", 'tx_niit',              '$',    "TAXABLE LTCG TAX", 11),
        ("Total\nTax",         'total_tax',            '$',    "TOTALS",           13),
        ("Ord Income\nEff Rate", 'eff1',                'pct',  "TOTALS",            9),
        ("All-In\nEff Rate",   'eff2',                 'pct',  "TOTALS",            9),
        ("MAGI",               'magi',                 '$',    "TOTALS",           14),
        ("After-Tax\nCash",    'after_tax_cash',       '$',    "NET CASH",         14),
        ("RMD\nRequired",      'rmd',                  '$',    "RMD",              12),
        ("Insurance\n(incl IRMAA)", 'ins',             '$',    "INSURANCE",        14),
        ("IRMAA\nSurcharge",   'irmaa',                '$',    "INSURANCE",        12),
        ("ACA\nHH Size",  'aca_hh_size',          'int',  "ACA",               8),
        ("ACA Subsidy\nCliff (400% FPL)", 'aca_cliff', '$',    "ACA",              16),
    ]

    GROUP_COLORS = {
        "YEAR":BLUE_H, "YEAR-END BALANCES":TEAL_H, "WITHDRAWALS":GREEN_H,
        "ORDINARY TAX":RED_H, "DIVIDEND TAX":"880E4F", "TAXABLE LTCG TAX":"BF360C",
        "TOTALS":"212121", "NET CASH":BROWN_H, "RMD":PURPLE_H, "INSURANCE":BROWN_H,
        "ACA":"00838F",
    }

    fallback_note = f" ⚠ Tax Law {p['tax_law_year']} not available — using {actual_yr}" if fallback_used else ""
    if p.get('gogo_rate', 0) > 0:
        gogo_yrs   = p.get('gogo_years',   0) or p['retire_years'] // 3
        slowgo_yrs = p.get('slowgo_years', 0) or p['retire_years'] // 3
        rate_str = (f"Go-Go {p['gogo_rate']*100:.0f}% (yrs 1–{gogo_yrs}) → "
                    f"Slow-Go {p['slowgo_rate']*100:.0f}% (yrs {gogo_yrs+1}–{gogo_yrs+slowgo_yrs}) → "
                    f"No-Go {p['nogo_rate']*100:.0f}% (yrs {gogo_yrs+slowgo_yrs+1}–{p['retire_years']})")
    else:
        rate_str = f"Gross Withdrawal Rate: {p.get('withdraw_rate', 0)*100:.0f}%"
    last_end_port = rows[-1]['end_port']
    title = (f"RETIREMENT WITHDRAWAL PLANNER — ANNUAL SUMMARY"
             f" | {rate_str}"
             f" | Roth Conversion Bracket: {p['conv_bracket']*100:.0f}%"
             f" | Tax Law: {actual_yr}{fallback_note}"
             f" | {p['retire_year']}–{p['retire_year']+p['retire_years']-1}"
             f" | Ending Portfolio Balance: ${last_end_port:,.0f}")
    t = ws.cell(1, 1, title)
    t.font = Font(bold=True, size=12, color=BLUE_H if not fallback_used else RED_H)
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")

    group_spans = {}
    for ci, (_,_,_,group,_) in enumerate(COLS,1):
        if group not in group_spans: group_spans[group]=[ci,ci]
        else: group_spans[group][1]=ci

    for group,(start,end) in group_spans.items():
        c=ws.cell(3,start,group); c.fill=hfill(GROUP_COLORS[group])
        c.font=Font(bold=True,color=WHITE,size=9)
        c.alignment=Alignment(horizontal='center',vertical='center')
        if end>start: ws.merge_cells(f"{get_column_letter(start)}3:{get_column_letter(end)}3")

    for ci,(hdr,_,_,group,w) in enumerate(COLS,1):
        c=ws.cell(4,ci,hdr); c.fill=hfill(GROUP_COLORS[group])
        c.font=Font(bold=True,color=WHITE,size=9)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[4].height=32

    for ri,row in enumerate(rows,5):
        for ci,(_,key,fmt,group,_) in enumerate(COLS,1):
            val=row.get(key,0)
            c=ws.cell(ri,ci,val)
            c.alignment=Alignment(horizontal='center' if fmt=='int' else 'right',vertical='center')
            c.border=tborder()
            if fmt=='$':   c.number_format='#,##0'
            if fmt=='pct': c.number_format='0.0%'
            if fmt=='int': c.number_format='0'
            if key=='phase':
                c.alignment=Alignment(horizontal='center',vertical='center')
                c.font=Font(bold=True,size=8,color=WHITE)
                c.fill=cfill("2E7D32") if val=='Go-Go' else cfill("E65100") if val=='Slow-Go' else cfill("B71C1C")
            elif key=='hb_fed':
                c.fill=cfill("FFF9C4") if val<=0.22 else cfill("FFE0B2") if val<=0.24 else cfill("FFCDD2")
            elif key=='total_tax': c.fill=cfill("FFEBEE")
            elif key=='rmd' and val>0: c.fill=cfill("EDE7F6"); c.font=Font(bold=True,size=9)
            elif key=='irmaa':
                if row['s1_age']<65 and row['s2_age']<65:
                    c.value="N/A"; c.fill=cfill("F5F5F5"); c.font=Font(italic=True,color="AAAAAA",size=8)
                elif val>0:
                    c.fill=cfill("FFF8E1"); c.font=Font(italic=True,color="888888",size=9)
            elif key == 'aca_hh_size':
                if val == 0:
                    c.value = "—"
                    c.fill = cfill("E8F5E9")
                    c.font = Font(italic=True, color="AAAAAA", size=9)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    c.fill = cfill("E3F2FD")
                    c.font = Font(bold=True, size=10)
                    c.alignment = Alignment(horizontal='center', vertical='center')
            elif key == 'aca_cliff':
                if val == 0.0:
                    c.value = "Medicare"
                    c.fill = cfill("E8F5E9")
                    c.font = Font(italic=True, color="555555", size=9)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                elif row['magi'] > val:
                    c.fill = cfill("FFCDD2")   # red — MAGI above cliff, no subsidies
                    c.font = Font(size=9, color="B71C1C")
                else:
                    c.fill = cfill("E8F5E9")   # green — eligible for subsidies
                    c.font = Font(size=9, color="1B5E20")
            elif key in ('trad_open','roth_open','taxable_open'): c.fill=cfill("E3F2FD")
            elif key in ('tx_fed','tx_or','tx_niit') and val==0:
                c.fill=cfill("F5F5F5"); c.font=Font(color="CCCCCC",size=8)
            elif ri%2==0: c.fill=cfill("FAFAFA")

    tr=len(rows)+5
    ws.cell(tr,1,"TOTALS").fill=hfill(BLUE_H)
    ws.cell(tr,1).font=Font(bold=True,color=WHITE,size=10)
    sum_keys={'trad_wd','taxable_principal_wd','roth_wd','div','roth_conv','ss1','ss2',
              'gross_target','f_tax_ord','or_tax_ord','div_fed','div_or','div_niit',
              'tx_fed','tx_or','tx_niit','fed_taxable_income','total_tax',
              'after_tax_cash','ins','rmd'}
    for ci,(_,key,_,_,_) in enumerate(COLS,1):
        if key in sum_keys:
            c=ws.cell(tr,ci,sum(r[key] for r in rows))
            c.number_format='#,##0'; c.font=Font(bold=True,size=9)
            c.fill=cfill("E8EAF6"); c.border=tborder()


def write_ss_scenarios(wb, p):
    if "SS Scenarios" in wb.sheetnames: del wb["SS Scenarios"]
    ws = wb.create_sheet("SS Scenarios")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = RED_H
    t=ws.cell(1,1,"SOCIAL SECURITY CLAIMING SCENARIOS — Lifetime Comparison")
    t.font=Font(bold=True,size=12,color=BLUE_H); ws.merge_cells("A1:H1")
    hdrs=["Scenario","S1 Claims","S2 Claims","S1 Starts","S2 Starts","Total SS","S1 Total","S2 Total"]
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(3,ci,h); c.fill=hfill(BLUE_H); c.font=Font(bold=True,color=WHITE)
        c.alignment=Alignment(horizontal='center')
    for ri,(name,p_age,w_age) in enumerate([("A — Early Both",62,62),("B — FRA Both",67,67),("C — Delay Primary ✓",70,67)],4):
        pia1=p['ss1_fra_mo']*12; pia2=p['ss2_fra_mo']*12
        p_start=p['spouse1_birth']+p_age; w_start=p['spouse2_birth']+w_age
        p_base=pia1*SS_CLAIMING_PCT[p_age]
        w_base=max(pia2*SS_CLAIMING_PCT.get(w_age,1.0),pia1*0.50)
        s1_tot=s2_tot=0.0
        for i in range(p['retire_years']):
            yr=p['retire_year']+i
            s1_tot+=p_base*(1+p['ss_cola'])**(yr-p_start) if yr>=p_start else 0
            s2_tot+=w_base*(1+p['ss_cola'])**(yr-w_start) if yr>=w_start else 0
        for ci,val in enumerate([name,p_age,w_age,p_start,w_start,s1_tot+s2_tot,s1_tot,s2_tot],1):
            c=ws.cell(ri,ci,val); c.alignment=Alignment(horizontal='center'); c.border=tborder()
            if isinstance(val,float): c.number_format='#,##0'
            if ri==6: c.fill=cfill("E8F5E9"); c.font=Font(bold=True)
    note=ws.cell(7,1,"✓ Recommended: Delay Spouse 1 to 70. Multi-year Roth conversion window. Max survivor benefit.")
    note.font=Font(italic=True,size=9,color="444444"); ws.merge_cells("A7:H7")
    note.alignment=Alignment(wrap_text=True); ws.row_dimensions[7].height=30
    for ci,w in enumerate([28,10,10,10,10,16,14,14],1):
        ws.column_dimensions[get_column_letter(ci)].width=w


def write_insurance_detail(wb, rows, p):
    if "Insurance Detail" in wb.sheetnames: del wb["Insurance Detail"]
    ws = wb.create_sheet("Insurance Detail")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00695C"
    t=ws.cell(1,1,"HEALTH INSURANCE DETAIL — ACA Gold (pre-65) & Medicare (post-65)")
    t.font=Font(bold=True,size=12,color=TEAL_H); ws.merge_cells("A1:H1")
    hdrs=["Year","S1 Age","S2 Age","S1 Coverage","S1 Annual","S2 Coverage","S2 Annual",
          "Children on ACA","Children Cost","Base Insurance","IRMAA Surcharge","Total Annual"]
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(3,ci,h); c.fill=hfill(TEAL_H); c.font=Font(bold=True,color=WHITE,size=9)
        c.alignment=Alignment(horizontal='center',wrap_text=True)
    ws.row_dimensions[3].height=28
    for ri,row in enumerate(rows,4):
        yr=row['year']; s1_age=row['s1_age']; s2_age=row['s2_age']; yrs=yr-2026
        s1_cov="Medicare" if s1_age>=65 else "ACA Gold"
        s2_cov="Medicare" if s2_age>=65 else "ACA Gold"
        aca_mo_base = p['aca_base_21']*(1+p['aca_inf'])**yrs
        s1_cost=p['medicare_base_mo']*12*(1+p['med_inf'])**yrs if s1_age>=65 else aca_mo_base*aca_age_factor(s1_age)*12
        s2_cost=p['medicare_base_mo']*12*(1+p['med_inf'])**yrs if s2_age>=65 else aca_mo_base*aca_age_factor(s2_age)*12
        kids = children_on_aca(yr, p)
        kids_cost = sum(aca_mo_base*aca_age_factor(max(a,21))*12 for a in kids)
        kids_desc = ", ".join(f"age {a}" for a in kids) if kids else "—"
        base_ins = s1_cost+s2_cost+kids_cost
        # IRMAA: read from simulation rows dict if available
        row_irmaa = next((r['irmaa'] for r in rows if r['year']==yr), 0.0)
        data=[yr,s1_age,s2_age,s1_cov,s1_cost,s2_cov,s2_cost,kids_desc,kids_cost,
              base_ins,row_irmaa,base_ins+row_irmaa]
        for ci,val in enumerate(data,1):
            c=ws.cell(ri,ci,val); c.alignment=Alignment(horizontal='center'); c.border=tborder()
            if isinstance(val,float): c.number_format='#,##0'
            if ci in(4,5) and s1_cov=="Medicare": c.fill=cfill("E8F5E9")
            if ci in(6,7) and s2_cov=="Medicare": c.fill=cfill("E8F5E9")
            if ci in(8,9) and kids: c.fill=cfill("FFF8E1")
            if ci==11 and isinstance(val,float) and val>0:
                c.fill=cfill("FFF3E0"); c.font=Font(bold=True,color="E65100",size=9)
            elif ri%2==0 and not(ci in(4,5) and s1_cov=="Medicare") \
                         and not(ci in(6,7) and s2_cov=="Medicare") \
                         and not(ci in(8,9) and kids):
                c.fill=cfill("F5F5F5")
    for ci,w in enumerate([7,7,7,12,14,12,14,18,14,14,14,16],1):
        ws.column_dimensions[get_column_letter(ci)].width=w


def write_rmd_projections(wb, rows, p, tl):
    if "RMD Projections" in wb.sheetnames: del wb["RMD Projections"]
    ws = wb.create_sheet("RMD Projections")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PURPLE_H
    rmd_start_yr=p['spouse1_birth']+tl.get('rmd_start_age',75)
    t=ws.cell(1,1,f"RMD PROJECTIONS — Spouse 1 (born {p['spouse1_birth']}, RMD starts age {tl.get('rmd_start_age',75)} = {rmd_start_yr})")
    t.font=Font(bold=True,size=12,color=PURPLE_H); ws.merge_cells("A1:G1")
    ws.cell(2,1,f"QCDs available from age 70½ ({p['spouse1_birth']+71}). ${tl['qcd_limit']:,}/person/yr. Reduces IRMAA MAGI.")
    ws.cell(2,1).font=Font(italic=True,size=9,color="555555"); ws.merge_cells("A2:G2")
    hdrs=["Year","S1 Age","Trad IRA Balance\n(opening)","IRS ULT Factor","RMD Required","Planned Trad W/D","Status"]
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(4,ci,h); c.fill=hfill(PURPLE_H); c.font=Font(bold=True,color=WHITE,size=9)
        c.alignment=Alignment(horizontal='center',wrap_text=True)
    ws.row_dimensions[4].height=30
    ri=5
    for row in rows:
        if row['s1_age']<70: continue
        rmd_age=tl.get('rmd_start_age',75)
        factor=RMD_FACTORS.get(min(row['s1_age'],98),"N/A") if row['s1_age']>=rmd_age else "Pre-RMD"
        rmd=row['rmd']; planned=row['trad_wd']
        status="⚠ UNDER" if isinstance(rmd,float) and rmd>0 and planned<rmd else("✓ OK" if rmd>0 else "—")
        for ci,val in enumerate([row['year'],row['s1_age'],row['trad_open'],factor,rmd,planned,status],1):
            c=ws.cell(ri,ci,val); c.alignment=Alignment(horizontal='center'); c.border=tborder()
            if isinstance(val,float) and val>100: c.number_format='#,##0'
            if val=="⚠ UNDER": c.fill=cfill("FFCDD2"); c.font=Font(bold=True,color="B71C1C")
        ri+=1
    for ci,w in enumerate([7,7,18,12,14,16,12],1):
        ws.column_dimensions[get_column_letter(ci)].width=w


def write_strategy_summary(wb, rows, p, actual_yr, fallback_used):
    if "Strategy Summary" in wb.sheetnames: del wb["Strategy Summary"]
    ws = wb.create_sheet("Strategy Summary")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREY_H

    def section(row,title,color):
        c=ws.cell(row,1,title); c.font=Font(bold=True,size=12,color=color)
        c.fill=cfill("F5F5F5"); ws.merge_cells(f"A{row}:C{row}"); return row+1

    def bullet(row,text,sub=False):
        c=ws.cell(row,1,("    • " if sub else "• ")+text)
        c.font=Font(size=10 if not sub else 9,color="222222" if not sub else "555555")
        c.alignment=Alignment(wrap_text=True); ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height=18; return row+1

    t=ws.cell(1,1,"RETIREMENT STRATEGY SUMMARY")
    t.font=Font(bold=True,size=14,color=BLUE_H); ws.merge_cells("A1:C1")
    if fallback_used:
        ws.cell(2,1,f"⚠  Tax Law Year {p['tax_law_year']} not in database — using {actual_yr} (latest available)")
        ws.cell(2,1).font=Font(bold=True,size=10,color=RED_H); ws.merge_cells("A2:C2")
    r=3

    r=section(r,"WITHDRAWAL ORDER & STRATEGY",BLUE_H)
    if p.get('gogo_rate', 0) > 0:
        r=bullet(r,f"3-Phase withdrawal: Go-Go {p['gogo_rate']*100:.0f}% → Slow-Go {p['slowgo_rate']*100:.0f}% → No-Go {p['nogo_rate']*100:.0f}%")
    else:
        r=bullet(r,f"{p.get('withdraw_rate',0)*100:.0f}% gross withdrawal rate on total portfolio each year")
    r=bullet(r,"Step 1 — Tax-Deferred: up to the Roth Conversion bracket ceiling")
    r=bullet(r,"Step 2 — Taxable Brokerage: principal fills gap (Fed LTCG + OR 9.9% + NIIT)",True)
    r=bullet(r,"Step 3 — Roth IRA: last resort, tax-free")
    r=bullet(r,f"Roth Conversion: fill remaining {p['conv_bracket']*100:.0f}% bracket headroom each year")

    r+=1; r=section(r,"SOCIAL SECURITY",RED_H)
    s1_ben=p['ss1_fra_mo']*SS_CLAIMING_PCT[p['ss1_claim_age']]*12
    s2_ben=max(p['ss2_fra_mo'],p['ss1_fra_mo']*0.50)*SS_CLAIMING_PCT[p['ss2_claim_age']]*12
    r=bullet(r,f"Spouse 1: claims at {p['ss1_claim_age']} ({p['spouse1_birth']+p['ss1_claim_age']}) → ${s1_ben:,.0f}/yr")
    r=bullet(r,f"Spouse 2: claims at {p['ss2_claim_age']} ({p['spouse2_birth']+p['ss2_claim_age']}) → ${s2_ben:,.0f}/yr")

    r+=1; r=section(r,"HEALTH INSURANCE TIMELINE",TEAL_H)
    r=bullet(r,f"Both on ACA Gold until Spouse 1 turns 65 ({p['spouse1_birth']+65})")
    r=bullet(r,f"Spouse 1 → Medicare {p['spouse1_birth']+65}; Spouse 2 → Medicare {p['spouse2_birth']+65}")
    r=bullet(r,"IRMAA shown only for Medicare-eligible years (2-yr MAGI lookback)")

    r+=1; r=section(r,"TAX LAW NOTES",ORANGE_H)
    r=bullet(r,f"Tax law base year: {actual_yr}" + (f" (⚠ {p['tax_law_year']} requested but not available)" if fallback_used else ""))
    r=bullet(r,f"All brackets inflated {p['inflation']*100:.1f}%/yr from {actual_yr} forward")
    r=bullet(r,"Oregon taxes ALL retirement income incl. SS — no preferential LTCG rate")
    r=bullet(r,"See Tax Tables tab for exact bracket values used in this run",True)

    ws.column_dimensions['A'].width=80


def write_monte_carlo(wb, rows, p):
    import random, math
    if "Monte Carlo" in wb.sheetnames: del wb["Monte Carlo"]
    ws = wb.create_sheet("Monte Carlo")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BROWN_H

    N_SIM   = 2_000
    MU      = 0.075   # log-normal mean annual return
    SIGMA   = 0.12    # log-normal std dev
    YEARS   = p['retire_years']
    random.seed(42)

    # Net portfolio withdrawal each year = gross draws (taxes paid FROM these, not in addition).
    # Dividends are generated by the taxable account and taken as cash (reduce balance).
    # SS is external income — does NOT come from the portfolio.
    net_wds = [r['trad_wd'] + r['roth_wd'] + r['taxable_principal_wd'] + r['div'] for r in rows]

    start_port = p['trad_bal'] + p['roth_bal'] + p['taxable_bal']

    def run_one(returns):
        bal = start_port
        for i, ret in enumerate(returns):
            bal = max(0, (bal - net_wds[i]) * (1 + ret))
        return bal

    # Generate all simulations
    log_mu    = math.log(1 + MU) - 0.5 * SIGMA**2
    all_paths = []
    for _ in range(N_SIM):
        rets = [math.exp(random.gauss(log_mu, SIGMA)) - 1 for _ in range(YEARS)]
        all_paths.append(rets)

    # Year-by-year portfolio balances for each sim
    year_balances = []   # year_balances[yr_idx] = sorted list of balances across sims
    for yr_idx in range(YEARS):
        bals = []
        for rets in all_paths:
            bal = start_port
            for i in range(yr_idx + 1):
                bal = max(0, (bal - net_wds[i]) * (1 + rets[i]))
            bals.append(bal)
        year_balances.append(sorted(bals))

    def percentile(sorted_vals, pct):
        idx = int(len(sorted_vals) * pct / 100)
        return sorted_vals[min(idx, len(sorted_vals)-1)]

    # Bad-sequence scenario: -35% crash in year 3
    bad_base = [MU] * YEARS
    bad_base[2] = -0.35
    bad_path_bals = []
    bal = start_port
    for i in range(YEARS):
        bal = max(0, (bal - net_wds[i]) * (1 + bad_base[i]))
        bad_path_bals.append(bal)

    # Success rate: sims where final balance > 0
    final_bals = [run_one(rets) for rets in all_paths]
    success_rate = sum(1 for b in final_bals if b > 0) / N_SIM

    # ── Sheet layout ──
    t = ws.cell(1, 1, f"MONTE CARLO SIMULATION — {N_SIM:,} runs | μ={MU*100:.1f}% σ={SIGMA*100:.1f}% log-normal | "
                      f"Portfolio Success Rate: {success_rate*100:.1f}%")
    t.font = Font(bold=True, size=12, color=BROWN_H)
    ws.merge_cells("A1:H1")

    ws.cell(2, 1, f"Starting portfolio: ${start_port:,.0f} | "
                  f"Bad-sequence scenario: −35% crash in year 3, then {MU*100:.1f}%/yr | "
                  f"Withdrawals follow the deterministic spending path (inflation-adjusted).")
    ws.cell(2, 1).font = Font(italic=True, size=9, color="555555")
    ws.merge_cells("A2:H2")

    hdrs = ["Year", "Age S1", "Det. Portfolio\n(Base Case)",
            "P10\n(Worst 10%)", "P25", "P50\n(Median)", "P75", "P90\n(Best 10%)",
            "Bad Sequence\n(−35% Yr 3)"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(4, ci, h)
        c.fill = hfill(BROWN_H); c.font = Font(bold=True, color=WHITE, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = tborder()
    ws.row_dimensions[4].height = 30

    det_bals = []
    bal = start_port
    for i in range(YEARS):
        bal = max(0, (bal - net_wds[i]) * (1 + p['growth_rate']))
        det_bals.append(bal)

    PCTS = [10, 25, 50, 75, 90]
    for ri, i in enumerate(range(YEARS), 5):
        yr     = p['retire_year'] + i
        s1_age = p['spouse1_age'] + (yr - 2026)
        yb     = year_balances[i]

        row_vals = [yr, s1_age, det_bals[i]] + \
                   [percentile(yb, pc) for pc in PCTS] + \
                   [bad_path_bals[i]]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(ri, ci, val)
            c.border = tborder()
            c.alignment = Alignment(horizontal='right' if ci > 2 else 'center', vertical='center')
            if ci == 1: c.alignment = Alignment(horizontal='center')
            if isinstance(val, float) and val > 10:
                c.number_format = '#,##0'
            # Colour: red if balance = 0 (portfolio depleted)
            if ci > 2 and val == 0:
                c.fill = cfill("FFCDD2"); c.font = Font(color="B71C1C", size=9)
            elif ci == 6:   # P50 median — highlight
                c.fill = cfill("E8F5E9"); c.font = Font(bold=True, size=9)
            elif ci == 9:   # bad sequence
                c.fill = cfill("FFF3E0")
            elif ri % 2 == 0:
                c.fill = cfill("FAFAFA")

    # Summary row
    sr = 5 + YEARS
    ws.cell(sr, 1, "FINAL").fill = hfill(BROWN_H)
    ws.cell(sr, 1).font = Font(bold=True, color=WHITE)
    ws.cell(sr, 1).border = tborder()
    ws.cell(sr, 2, f"Yr {YEARS}").border = tborder()
    ws.cell(sr, 2).alignment = Alignment(horizontal='center')
    summary_vals = [det_bals[-1]] + [percentile(year_balances[-1], pc) for pc in PCTS] + [bad_path_bals[-1]]
    for ci, val in enumerate(summary_vals, 3):
        c = ws.cell(sr, ci, val)
        c.number_format = '#,##0'; c.border = tborder()
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal='right')
        c.fill = cfill("FFCDD2") if val == 0 else cfill("C8E6C9")
    ws.row_dimensions[sr].height = 20

    # Success rate note
    nr = sr + 2
    note = ws.cell(nr, 1,
        f"Portfolio Success Rate: {success_rate*100:.1f}% of {N_SIM:,} simulations end with balance > $0  |  "
        f"P10=worst 10%, P50=median outcome, P90=best 10%  |  "
        f"Bad-sequence: −35% crash year 3, then {MU*100:.1f}%/yr thereafter.")
    note.font = Font(italic=True, size=9, color="444444")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{nr}:I{nr}")
    ws.row_dimensions[nr].height = 24

    col_widths = [7, 7, 16, 14, 12, 14, 12, 14, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_roth_conversion_comparison(wb, p, tl, base_yr):
    sheet_name = "Roth Conversion Scenario Comparison"
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PURPLE_H

    scenarios = [
        ("No Conversion",  {**p, 'conv_bracket': 0}),
        ("Convert to 22%", {**p, 'conv_bracket': 0.22}),
        ("Convert to 24%", {**p, 'conv_bracket': 0.24}),
    ]

    # Run all 3 simulations
    sim_results = [(label, run_simulation(sp, tl, base_yr)) for label, sp in scenarios]

    t = ws.cell(1, 1, "ROTH CONVERSION SCENARIO COMPARISON — No Conversion vs 22% vs 24%")
    t.font = Font(bold=True, size=13, color=PURPLE_H)
    ws.merge_cells(f"A1:{get_column_letter(2 + len(scenarios)*2)}1")

    # Header row
    hdrs = ["Year", "Age S1/S2"] + [f"{lbl}\nNet Spendable" for lbl, _ in sim_results] \
         + [f"{lbl}\nRoth Conv" for lbl, _ in sim_results]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(3, ci, h)
        c.fill = hfill(PURPLE_H); c.font = Font(bold=True, color=WHITE, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = tborder()
    ws.row_dimensions[3].height = 30

    n = len(sim_results)
    for ri, i in enumerate(range(p['retire_years']), 4):
        yr = p['retire_year'] + i
        s1_age = p['spouse1_age'] + (yr - 2026)
        s2_age = p['spouse2_age'] + (yr - 2026)

        ws.cell(ri, 1, yr).border = tborder()
        ws.cell(ri, 2, f"{s1_age}/{s2_age}").border = tborder()
        ws.cell(ri, 2).alignment = Alignment(horizontal='center')

        spendables = [rows[i]['after_tax_net'] for _, rows in sim_results]
        convs      = [rows[i]['roth_conv']      for _, rows in sim_results]

        best_idx  = spendables.index(max(spendables))
        worst_idx = spendables.index(min(spendables))

        for si, val in enumerate(spendables):
            c = ws.cell(ri, 3 + si, val)
            c.number_format = '#,##0'; c.border = tborder()
            c.alignment = Alignment(horizontal='right', vertical='center')
            if si == best_idx:
                c.fill = cfill("C8E6C9"); c.font = Font(bold=True, color="1B5E20")
            elif si == worst_idx:
                c.fill = cfill("FFCDD2"); c.font = Font(bold=True, color="B71C1C")
            elif ri % 2 == 0:
                c.fill = cfill("F5F5F5")

        for si, val in enumerate(convs):
            c = ws.cell(ri, 3 + n + si, val)
            c.number_format = '#,##0'; c.border = tborder()
            c.alignment = Alignment(horizontal='right', vertical='center')
            if val > 0:
                c.fill = cfill("EDE7F6")
            elif ri % 2 == 0:
                c.fill = cfill("F5F5F5")

    # Totals row
    tr = 4 + p['retire_years']
    ws.cell(tr, 1, "TOTAL").fill = hfill(PURPLE_H)
    ws.cell(tr, 1).font = Font(bold=True, color=WHITE)
    ws.cell(tr, 1).border = tborder()
    ws.cell(tr, 2, "35-year sum").border = tborder()
    ws.cell(tr, 2).font = Font(italic=True, size=9, color="666666")

    total_spendables = [sum(rows[i]['after_tax_net'] for i in range(p['retire_years']))
                        for _, rows in sim_results]
    total_convs      = [sum(rows[i]['roth_conv']      for i in range(p['retire_years']))
                        for _, rows in sim_results]

    best_tot  = total_spendables.index(max(total_spendables))
    worst_tot = total_spendables.index(min(total_spendables))

    for si, val in enumerate(total_spendables):
        c = ws.cell(tr, 3 + si, val)
        c.number_format = '#,##0'; c.border = tborder()
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal='right', vertical='center')
        if si == best_tot:
            c.fill = cfill("A5D6A7"); c.font = Font(bold=True, size=11, color="1B5E20")
        elif si == worst_tot:
            c.fill = cfill("EF9A9A"); c.font = Font(bold=True, size=11, color="B71C1C")
        else:
            c.fill = cfill("EDE7F6")

    for si, val in enumerate(total_convs):
        c = ws.cell(tr, 3 + n + si, val)
        c.number_format = '#,##0'; c.border = tborder()
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.fill = cfill("EDE7F6")

    ws.row_dimensions[tr].height = 22

    # Legend note
    note_row = tr + 2
    note = ws.cell(note_row, 1,
        "Green = best total spendable   |   Red = worst total spendable   |"
        "   Purple = Roth conversion amount   |   All 3 scenarios use identical inputs except conversion bracket.")
    note.font = Font(italic=True, size=9, color="444444")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{note_row}:{get_column_letter(2 + n*2)}{note_row}")
    ws.row_dimensions[note_row].height = 24

    col_widths = [7, 10] + [16]*n + [14]*n
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else "Retirement_Planner.xlsx"

    if not os.path.exists(filepath):
        print(f"\nFile not found: '{filepath}'")
        print("Creating input template with default values...\n")
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        create_inputs_sheet(wb)
        wb.save(filepath)
        print(f"✓ Template created: {filepath}")
        print(f"\nNext steps:")
        print(f"  1. Open {filepath}")
        print(f"  2. Edit the amber cells in the 'Inputs' tab")
        print(f"     The orange 'Tax Law Year' cell controls which IRS tables are used.")
        print(f"     Available years: {sorted(TAX_LAW_DB.keys())}")
        print(f"  3. Save and run:  python3 build_retirement_planner.py {filepath}")
        return

    print(f"\nReading inputs from: {filepath}")
    p = read_inputs(filepath)

    tl, actual_yr, fallback_used = get_tax_law(p['tax_law_year'])
    if fallback_used:
        print(f"  ⚠  Tax law year {p['tax_law_year']} not in database — using {actual_yr} (latest available)")
        print(f"     Available years: {sorted(TAX_LAW_DB.keys())}")
    else:
        print(f"  Tax law year: {actual_yr}")

    print(f"  Portfolio: Trad ${p['trad_bal']:,.0f} | Taxable ${p['taxable_bal']:,.0f} | Roth ${p['roth_bal']:,.0f}")
    if p.get('gogo_rate', 0) > 0:
        print(f"  Withdrawal: Go-Go {p['gogo_rate']*100:.0f}% / Slow-Go {p['slowgo_rate']*100:.0f}% / No-Go {p['nogo_rate']*100:.0f}% | Conv Bracket: {p['conv_bracket']*100:.0f}% | {p['retire_year']}–{p['retire_year']+p['retire_years']-1}")
    else:
        print(f"  Withdrawal: {p.get('withdraw_rate',0)*100:.0f}% | Conv Bracket: {p['conv_bracket']*100:.0f}% | {p['retire_year']}–{p['retire_year']+p['retire_years']-1}")

    print(f"\nRunning simulation...")
    rows = run_simulation(p, tl, actual_yr)

    print("Writing output sheets...")
    wb = openpyxl.load_workbook(filepath)

    # Rebuild Inputs sheet with current values and correct row order
    create_inputs_sheet(wb, values=p)

    # Remove any sheets that are not Inputs and not in our current output set,
    # so stale sheets from old script versions don't persist.
    known_output_sheets = {
        "Tax Tables", "Assumptions", "Annual Summary", "SS Scenarios",
        "Roth Conversion Scenario Comparison", "Monte Carlo",
        "Insurance Detail", "RMD Projections", "Strategy Summary",
    }
    for name in list(wb.sheetnames):
        if name != "Inputs" and name not in known_output_sheets:
            del wb[name]
            print(f"  Removed stale sheet: {name}")

    write_assumptions_output(wb, rows, p, tl, actual_yr, actual_yr, fallback_used)
    write_annual_summary(wb, rows, p, actual_yr, fallback_used)
    write_ss_scenarios(wb, p)
    write_roth_conversion_comparison(wb, p, tl, actual_yr)
    write_monte_carlo(wb, rows, p)
    write_insurance_detail(wb, rows, p)
    write_rmd_projections(wb, rows, p, tl)
    write_strategy_summary(wb, rows, p, actual_yr, fallback_used)
    write_tax_tables(wb, tl, actual_yr, actual_yr, fallback_used, p)

    # Ensure tab order: Inputs first, Tax Tables second, then outputs
    sheet_order = ["Inputs", "Tax Tables", "Assumptions", "Annual Summary",
                   "SS Scenarios", "Roth Conversion Scenario Comparison", "Monte Carlo",
                   "Insurance Detail", "RMD Projections", "Strategy Summary"]
    for i, name in enumerate(sheet_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(filepath)

    total_at = sum(r['after_tax_cash'] for r in rows)
    last = rows[-1]
    end_roth = last['roth_open']*(1+p['growth_rate'])
    end_trad = last['trad_open']*(1+p['growth_rate'])
    end_tax  = last['taxable_open']*(1+p['price_return'])
    print(f"\n── RESULTS ─────────────────────────────────────────────")
    print(f"  Year 1 after-tax cash:      ${rows[0]['after_tax_cash']:>12,.0f}")
    print(f"  Average after-tax cash/yr:  ${sum(r['after_tax_cash'] for r in rows)/p['retire_years']:>12,.0f}")
    print(f"  Year {p['retire_years']} after-tax cash:     ${last['after_tax_cash']:>12,.0f}")
    print(f"  Total after-tax cash:       ${sum(r['after_tax_cash'] for r in rows):>12,.0f}")
    print(f"  Total taxes paid:      ${sum(r['total_tax'] for r in rows):>12,.0f}")
    print(f"  End portfolio:         ${end_trad+end_roth+end_tax:>12,.0f}  (Roth: ${end_roth:,.0f})")
    print(f"\n✓ Saved: {filepath}")
    print(f"  Tax Tables tab shows exact {actual_yr} law values used in this run.")


if __name__ == "__main__":
    main()
